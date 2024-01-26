from django.contrib import admin
from .models import WeatherEntry
import openpyxl
from django.http import HttpResponse

class WeatherEntryAdmin(admin.ModelAdmin):
    # Your existing configurations for the admin class

    def save_to_excel(self, request, queryset):
        # Your custom logic goes here
        # `queryset` contains the selected objects
        wb = openpyxl.Workbook()
        ws = wb.active

        headers = [field.name for field in WeatherEntry._meta.fields]
        for col_num, header in enumerate(headers, 1):
            col_letter = openpyxl.utils.get_column_letter(col_num)
            ws[f'{col_letter}'] = header

        # write data
        for row_num, entry in enumerate(queryset, 2):
            for col_name, field_name in enumerate(headers, 1):
                col_letter = openpyxl.utils.get_column_letter(col_num)
                ws[f'{col_letter}{row_num}'] = str(getattr(entry, field_name))

        # create response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=weather_entries.xlsx'


        # save to response
        wb.save(response)

        return response


    save_to_excel.short_description = "save_to_excel"

    actions = ['save_to_excel']
# Register the admin class with the model
admin.site.register(WeatherEntry, WeatherEntryAdmin)

